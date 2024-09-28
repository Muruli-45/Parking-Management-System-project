import sys
import openpyxl
from openpyxl import Workbook
import time
import csv
import os
from datetime import datetime

class Vehicle:
    def __init__(self, vehicle_number, vehicle_type, vehicle_name, owner_name):
        self.vehicle_number = vehicle_number
        self.vehicle_type = vehicle_type
        self.vehicle_name = vehicle_name
        self.owner_name = owner_name

    def __str__(self):
        return f"Vehicle Number: {self.vehicle_number}, Vehicle Type: {self.vehicle_type}, Vehicle Name: {self.vehicle_name}, Owner Name: {self.owner_name}"

class Reservation:
    def __init__(self, vehicle, date, time, fees):
        self.vehicle = vehicle
        self.date = date
        self.time = time
        self.fees = fees

    def __str__(self):
        return f"Vehicle Number: {self.vehicle.vehicle_number}, Date: {self.date}, Time: {self.time}, Fees: {self.fees}"

class Bill:
    def __init__(self, reservation, tax, penalty):
        self.reservation = reservation
        self.tax = tax
        self.penalty = penalty

    def calculate_total(self):
        return self.reservation.fees + self.tax + self.penalty

    def __str__(self):
        return f"Vehicle Number: {self.reservation.vehicle.vehicle_number}, Date: {self.reservation.date}, Time: {self.reservation.time}, Fees: {self.reservation.fees}, Tax: {self.tax}, Penalty: {self.penalty}, Total: {self.calculate_total()}"

class ParkingManagementSystem:
    def __init__(self):
        self.vehicles = []
        self.reservations = []
        self.bills = []
        self.bicycles = 30
        self.two_wheeler = 25
        self.four_wheeler = 10
        self.vehicle_numbers = []
        self.load_data()

    def load_data(self):
        try:
            with open("data.txt", "r") as file:
                for line in file:
                    values = line.strip().split(",")
                    if len(values) != 7:
                        print(f"Skipping line: {line.strip()}. Expected 7 values, got {len(values)}.")
                        continue
                    vehicle_number, vehicle_type, vehicle_name, owner_name, date, time, fees = values
                    vehicle = Vehicle(vehicle_number, vehicle_type, vehicle_name, owner_name)
                    reservation = Reservation(vehicle, date, time, int(fees))
                    self.vehicles.append(vehicle)
                    self.reservations.append(reservation)
                    self.vehicle_numbers.append(vehicle_number)
        except FileNotFoundError:
            print("File not found. Creating a new file.")
            open("data.txt", "w").close()

    def save_data(self):
        filename = "parking_data.xlsx"
        i = 1
        while os.path.exists(filename):
            filename = f"parking_data_{i}.xlsx"
            i += 1

        wb = Workbook()
        ws = wb.active
        ws.title = "Parking Data"

        ws['A1'] = 'Vehicle Number'
        ws['B1'] = 'Vehicle Type'
        ws['C1'] = 'Vehicle Name'
        ws['D1'] = 'Owner Name'
        ws['E1'] = 'Date'
        ws['F1'] = 'Time'
        ws['G1'] = 'Fees'
        ws['H1'] = 'Tax'
        ws['I1'] = 'Penalty'
        ws['J1'] = 'Total'

        row = 2
        for reservation in self.reservations:
            ws[f'A{row}'] = reservation.vehicle.vehicle_number
            ws[f'B{row}'] = reservation.vehicle.vehicle_type
            ws[f'C{row}'] = reservation.vehicle.vehicle_name
            ws[f'D{row}'] = reservation.vehicle.owner_name
            ws[f'E{row}'] = reservation.date
            ws[f'F{row}'] = reservation.time
            ws[f'G{row}'] = reservation.fees

            for bill in self.bills:
                if bill.reservation.vehicle.vehicle_number == reservation.vehicle.vehicle_number:
                    ws[f'H{row}'] = bill.tax
                    ws[f'I{row}'] = bill.penalty
                    ws[f'J{row}'] = bill.calculate_total()
                    break
            else:
                ws[f'H{row}'] = 0
                ws[f'I{row}'] = 0
                ws[f'J{row}'] = 0

            row += 1

        wb.save(filename)

    def add_vehicle(self, vehicle):
        self.vehicles.append(vehicle)

    def remove_vehicle(self, vehicle):
        self.vehicles.remove(vehicle)

    def add_reservation(self, reservation):
        self.reservations.append(reservation)

    def remove_reservation(self, reservation):
        self.reservations.remove(reservation)

    def add_bill(self, bill):
        self.bills.append(bill)

    def remove_bill(self, bill):
        self.bills.remove(bill)

    def display_menu(self):
        print("-----------------------------------------------")
        print("\t\tParking Management System")
        print("------------------------------------------------")
        print("1.Vehicle Entry")
        print("2.Remove Entry")
        print("3.View Parked Vehicle")
        print("4.View Left Parking Space")
        print("5.Amount Details")
        print("6.Bill")
        print("7.Update Vehicle Details")
        print("8.Close Programme")
        print("+---------------------------------------------+")

    def vehicle_entry(self ):
        no=True
        while no==True:
            v = input("\tEnter vehicle number (XXXX-XXXX) -").upper()
            if v=="":
                print("\tplease enter the vehicle number")
            elif v in self.vehicle_numbers:
                print("vehicle number already exist")
            elif len(v)==9:
                no=not True
                self.vehicle_numbers.append(v)
            else:
                print("######## enter valid vehicle no ")

        typee=True
        while typee==True:
            vehicle_type = str(input("\tEnter vehicle type(Bicycle=A/Two Wheeler=B/Four Wheeler=C):")).lower()
            if vehicle_type=="":
                print("Please enter any vehicle type first")
            elif vehicle_type == "a":
                vehicle_type = "Bicycle"
                if self.bicycles > 0:
                    self.bicycles -= 1
                else:
                    print("No Parking slots are available for Bicycles.")
                typee=not True
            elif vehicle_type == "b":
                vehicle_type = "Two Wheeler"
                if self.two_wheeler > 0:
                    self.two_wheeler -= 1
                else:
                    print("No parking slots are available for Two Wheelers")
                typee=not True
            elif vehicle_type == "c":
                vehicle_type = "Four Wheeler"
                if self.four_wheeler > 0:
                    self.four_wheeler -= 1  
                else:
                    print("No parking slots are available for Four Wheelers")
                typee=not True
            else:
                print("enter a valid option")

        vehicle_name=input("\tEnter vehicle name - ")
        owner_name=input("\tEnter owner name - ")

        d=True
        while d==True:
            date = input("\tEnter Date (DD-MM-YYYY) - ")
            if date=="":
                print("please enter the date")
            elif len(date)!=10:
                print("#### Please enter a valid date")
            else:
                d=not True

        t = True
        while t == True:
            time = input("\tEnter Time (HH:MM) - ")
            if time=="":
                print("###### Enter Time ######")
            elif len(time)!=5:
                print("###### Please Enter Valid Time ######")
            else:
                t=not True

        veh = Vehicle(v, vehicle_type, vehicle_name, owner_name)
        self.add_vehicle(veh)

        res = Reservation(veh, date, time, 0)
        self.add_reservation(res)

        print("\n..........................Record detail saved....................................")

    def remove_entry(self):
        v = input("\tEnter vehicle number to Delete(XXXX-XXXX) - ").upper()
        if v not in self.vehicle_numbers:
            print("\t---Vehicle number does not exist---")
        else:
            for vehicle in self.vehicles:
                if vehicle.vehicle_number == v:
                    if vehicle.vehicle_type == "Bicycle":
                        self.bicycles += 1
                    elif vehicle.vehicle_type == "Two Wheeler":
                        self.two_wheeler += 1
                    elif vehicle.vehicle_type == "Four Wheeler":
                        self.four_wheeler += 1
                self.remove_vehicle(vehicle)
                self.vehicle_numbers.remove(v)
                for reservation in self.reservations:
                    if reservation.vehicle.vehicle_number == v:
                        self.remove_reservation(reservation)
                print("\n.......................Removed Sucessfully............................")
                return
        print("###### No Such Entry ######")

    def view_parked_vehicle(self):
        count = 0
        print("--------------------------------------------------------------------------------------------------------------------")
        print("\t\t\t\tParked Vehicle")
        print("--------------------------------------------------------------------------------------------------------------------")
        print("Vehicle No.\tVehicle Type        Vehicle Name\t\t       Owner Name\t     Date\t\tTime")
        print ("--------------------------------------------------------------------------------------------------------------------")
        for reservation in self.reservations:
            count += 1
            print(reservation.vehicle.vehicle_number,"      ",reservation.vehicle.vehicle_type, "\t\t   ", reservation.vehicle.vehicle_name, "\t\t      ", reservation.vehicle.owner_name,"\t\t",reservation.date,"\t\t",reservation.time)
        print("----------------------------------------------------------------------------------------------------------")
        print("------------------------------------- Total Records - ", count, "-----------------------------------------")
        print("----------------------------------------------------------------------------------------------------------")

    def view_left_parking_space(self):
        print("-----------------------------------------------------------------------------------")
        print("\t\tSpaces Left For Parking")
        print("-----------------------------------------------------------------------------------")
        print("\tSpaces Available for Bicycle - ", self.bicycles)
        print("\tSpaces Available for Two Wheeler - ", self.two_wheeler)
        print("\tSpaces Available for Four Wheeler - ", self.four_wheeler)
        print("-----------------------------------------------------------------------------------")

    def amount_details(self):
        print("----------------------------------------------------------------------")
        print("\t\t  Parking Rate")
        print("----------------------------------------------------------------------")
        print("*1.Bicycle             Rs20/ Hour")
        print("*2.Two Wheeler         Rs40/ Hour")
        print("*3.Four Wheeler        Rs60/ Hour")
        print("---------------------------------------------------------------------")

    def generate_bill(self):
        vehicle_number = input("\tEnter vehicle number to Delete(XXXX-XXXX) - ").upper()
        for reservation in self.reservations:
            if reservation.vehicle.vehicle_number == vehicle_number:
                print("\tVehicle Check in time - ", reservation.time)
                print("\tVehicle Check in Date - ", reservation.date)
                print("\tVehicle Type - ", reservation.vehicle.vehicle_type)
                hours = int(input("\tEnter No. of Hours Vehicle Parked - "))
                if reservation.vehicle.vehicle_type == "Bicycle":
                    amount=hours*20
                elif reservation.vehicle.vehicle_type == "Two Wheeler":
                    amount=hours*40
                elif reservation.vehicle.vehicle_type == "Four Wheeler":
                    amount=hours*60
                else:
                    print("Invalid vehicle type")
                    return
                reservation.fees = amount  # Update the fees attribute
                tax=18/100*amount
                
                due_date = input("\tEnter due date (DD-MM-YYYY) - ")
                return_date = input("\tEnter return date (DD-MM-YYYY) - ")
                penalty = self.penalty(due_date, return_date)
                
                bill = Bill(reservation, tax, penalty)
                self.add_bill(bill)
                print("\t Parking Charge - ",amount)
                print("\tAdd. charge 18 % - ",tax)
                print("\tPenalty - ", penalty)
                print("\tTotal Charge - ", amount+tax+penalty)
                print("............................WELCOME...............................")  
                self.save_data()
                return
        print("###### No Such Entry ######")

    def penalty(self, due_date, return_date):
        due = datetime.strptime(due_date,"%d-%m-%Y")
        returned = datetime.strptime(return_date,"%d-%m-%Y")
        if returned <= due:
            return 0
        elif returned.year == due.year and returned.month == due.month:
            return 15 * (returned.day - due.day)
        elif returned.year == due.year:
            return 200
        else:
            return 600
    
    def update_entry(self):
        v = input("\tEnter vehicle number to Update(XXXX-XXXX) - ").upper()
        for vehicle in self.vehicles:
            if vehicle.vehicle_number == v:
                print("\n-------------------Update Options-------------------")
                print("1. Update Vehicle Name")
                print("2. Update Owner Name")
                print("3. Update Date")
                print("4. Update Time")
                try:
                    choice = int(input("\tSelect option: "))
                except ValueError:
                    print("Invalid input. Please enter a number.")
                    return

                if choice == 1:
                    vehicle_name = input("\tEnter new vehicle name: ")
                    vehicle.vehicle_name = vehicle_name
                    for reservation in self.reservations:
                        if reservation.vehicle.vehicle_number == v:
                            reservation.vehicle.vehicle_name = vehicle_name
                elif choice == 2:
                    owner_name = input("\tEnter new owner name: ")
                    vehicle.owner_name = owner_name
                    for reservation in self.reservations:
                        if reservation.vehicle.vehicle_number == v:
                            reservation.vehicle.owner_name = owner_name
                elif choice == 3:
                    date = input("\tEnter new date (DD-MM-YYYY): ")
                    for reservation in self.reservations:
                        if reservation.vehicle.vehicle_number == v:
                            reservation.date = date
                elif choice == 4:
                    time = input("\tEnter new time (HH:MM): ")
                    for reservation in self.reservations:
                        if reservation.vehicle.vehicle_number == v:
                            reservation.time = time
                else:
                    print("---------------INVALID CHOICE---------------")
                print("\n-------------------Updated Successfully-------------------")
                return
        print("###### No Such Entry ######")
    def run(self):
        while True:
            self.display_menu()
            try:
                ch=int(input("\tSelect option:"))
            except ValueError:
                print("Invalid input.Please enter a number.")
                continue

            if ch==1:
                self.vehicle_entry()
                self.save_data()
            elif ch == 2:
                self.remove_entry()
                self.save_data()
            elif ch == 3:
                self.view_parked_vehicle()
            elif ch == 4:
                self.view_left_parking_space()
            elif ch == 5:
                self.amount_details()
            elif ch == 6:
                self.generate_bill()
            elif ch == 7:
                self.update_entry()
            elif ch == 8:
                self.save_data()
                print("...........................Thank you for using our service......................................")
                print("                         **********(-- Bye Bye --)**********")
                break
            else:
                print("---------------INVALID CHOICE---------------")
if __name__ == "__main__":
    pm= ParkingManagementSystem()
    pm.run()